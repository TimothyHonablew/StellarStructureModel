#Name: Timothy Honablew
#Date: November 13th, 2023 CE
#Purpose: Holding the appropriate opacity table for my stellar calculation and calling the proper interpolation function.

from tkinter import Tk
from tkinter.filedialog import askopenfilename

from openpyxl import load_workbook

import numpy as np
from scipy.interpolate import LinearNDInterpolator

#Opening the excel file that contaitns the opacity table
Tk().withdraw();
fileName = askopenfilename();

#Accessing the data in the excel file
wb = load_workbook(fileName);
ws = wb.active;

columnA = ws['A'];
columnB = ws['B'];
columnC = ws['C'];
columnD = ws['D'];
columnE = ws['E'];
columnF = ws['F'];
columnG = ws['G'];
columnH = ws['H'];
columnI = ws['I'];
columnJ = ws['J'];
columnK = ws['K'];
columnL = ws['L'];
columnM = ws['M'];
columnN = ws['N'];
columnO = ws['O'];
columnP = ws['P'];
columnQ = ws['Q'];
columnR = ws['R'];
columnS = ws['S'];
columnT = ws['T'];

#Putting the data into a usable format
A = [cell.value for cell in columnA[2:]];
B = [cell.value for cell in columnB[2:]];
C = [cell.value for cell in columnC[2:]];
D = [cell.value for cell in columnD[2:]];
E = [cell.value for cell in columnE[2:]];
F = [cell.value for cell in columnF[2:]];
G = [cell.value for cell in columnG[2:]];
H = [cell.value for cell in columnH[2:]];
I = [cell.value for cell in columnI[2:]];
J = [cell.value for cell in columnJ[2:]];
K = [cell.value for cell in columnK[2:]];
L = [cell.value for cell in columnL[2:]];
M = [cell.value for cell in columnM[2:]];
N = [cell.value for cell in columnN[2:]];
O = [cell.value for cell in columnO[2:]];
P = [cell.value for cell in columnP[2:]];
Q = [cell.value for cell in columnQ[2:]];
R = [cell.value for cell in columnR[2:]];
S = [cell.value for cell in columnS[2:]];
T = [cell.value for cell in columnT[2:]];

logR = [-8.0, -7.5, -7.0, -6.5, -6.0, -5.5, -5.0, -4.5, -4.0, -3.5, -3.0, -2.5, -2.0, -1.5, -1.0, -0.5, 0.0, 0.5, 1.0];


#Creating and filling the single arrays to hold all the data so that the 2D interpolation funciton can be called
temp = [];
rad = [];
opa = [];

arrays = [B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q, R, S, T];

for i in range(len(arrays)):
    for j in range(len(arrays[i])):
        temp.append(A[j]);
        rad.append(logR[i]);
        opa.append(arrays[i][j]);


#Calling the 2D interpolation
opacityFunc = LinearNDInterpolator(list(zip(temp, rad)), opa);

def currentOpacity(temp, rad):
    return opacityFunc(temp, rad);